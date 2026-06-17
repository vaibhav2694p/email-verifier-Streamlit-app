from __future__ import annotations

import io
from datetime import datetime

import pandas as pd
from fpdf import FPDF
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


GREEN_FILL = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
RED_FILL = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _status_fill(val: str) -> PatternFill:
    if val == "VALID":
        return GREEN_FILL
    elif val == "RISKY":
        return YELLOW_FILL
    elif val == "INVALID":
        return RED_FILL
    return PatternFill()


def _apply_column_widths(ws, df: pd.DataFrame) -> None:
    for i, col in enumerate(df.columns, 1):
        max_len = max(
            df[col].astype(str).map(len).max() if len(df) else 0,
            len(str(col)),
        )
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 3, 40)


def _style_header(ws, df: pd.DataFrame) -> None:
    for i, col in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def _color_rows(ws, df: pd.DataFrame, status_col: str) -> None:
    for row_idx in range(2, len(df) + 2):
        status_val = str(ws.cell(row=row_idx, column=df.columns.get_loc(status_col) + 1).value)
        fill = _status_fill(status_val)
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)


def dataframe_to_csv_bytes(df: pd.DataFrame) -> bytes:
    cols = list(df.columns)
    rename = {"Email Address": "Email Address"}
    export = df.rename(columns=rename)
    return export.to_csv(index=False).encode("utf-8")


def dataframe_to_xlsx_bytes(df: pd.DataFrame) -> bytes:
    """Full report with Verification Results + Summary sheets, color-coded."""
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        results_sheet = df.copy()

        results_sheet.to_excel(writer, index=False, sheet_name="Verification Results")

        ws = writer.sheets["Verification Results"]
        _style_header(ws, results_sheet)
        _apply_column_widths(ws, results_sheet)
        _color_rows(ws, results_sheet, "Overall Status")

        _build_summary_sheet(writer, df)

        _build_separate_sheets(writer, df)

    return output.getvalue()


def _build_summary_sheet(writer: pd.ExcelWriter, df: pd.DataFrame) -> None:
    total = len(df)
    valid_count = int((df["Overall Status"] == "VALID").sum())
    risky_count = int((df["Overall Status"] == "RISKY").sum())
    invalid_count = int((df["Overall Status"] == "INVALID").sum())

    summary_data = {
        "Metric": [
            "Total Emails",
            "Valid Emails",
            "Risky Emails",
            "Invalid Emails",
            "",
            "Generated",
        ],
        "Value": [
            total,
            valid_count,
            risky_count,
            invalid_count,
            "",
            datetime.now().strftime("%d-%b-%Y %I:%M %p"),
        ],
    }
    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(writer, index=False, sheet_name="Summary")

    ws = writer.sheets["Summary"]
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15

    for i, col in enumerate(summary_df.columns, 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    fills = [GREEN_FILL, YELLOW_FILL, RED_FILL]
    status_rows = [2, 3, 4]
    for idx, row_num in enumerate(status_rows):
        for col in range(1, 3):
            ws.cell(row=row_num, column=col).fill = fills[idx]
            ws.cell(row=row_num, column=col).border = THIN_BORDER

    ws.cell(row=7, column=1).value = "Color Coding:"
    ws.cell(row=7, column=1).font = Font(bold=True)
    ws.cell(row=8, column=1).value = "🟢 Green  = Valid Email"
    ws.cell(row=8, column=1).fill = GREEN_FILL
    ws.cell(row=9, column=1).value = "🟡 Yellow = Risky Email"
    ws.cell(row=9, column=1).fill = YELLOW_FILL
    ws.cell(row=10, column=1).value = "🔴 Red    = Invalid Email"
    ws.cell(row=10, column=1).fill = RED_FILL


def _build_separate_sheets(writer: pd.ExcelWriter, df: pd.DataFrame) -> None:
    categories = {
        "Valid_Emails": df[df["Overall Status"] == "VALID"],
        "Risky_Emails": df[df["Overall Status"] == "RISKY"],
        "Invalid_Emails": df[df["Overall Status"] == "INVALID"],
    }

    for sheet_name, subset_df in categories.items():
        if subset_df.empty:
            continue
        subset_df.to_excel(writer, index=False, sheet_name=sheet_name)

        ws = writer.sheets[sheet_name]
        _style_header(ws, subset_df)
        _apply_column_widths(ws, subset_df)
        _color_rows(ws, subset_df, "Overall Status")


def dataframe_valid_xlsx_bytes(df: pd.DataFrame) -> bytes:
    return _filtered_xlsx(df, "VALID")


def dataframe_risky_xlsx_bytes(df: pd.DataFrame) -> bytes:
    return _filtered_xlsx(df, "RISKY")


def dataframe_invalid_xlsx_bytes(df: pd.DataFrame) -> bytes:
    return _filtered_xlsx(df, "INVALID")


def _filtered_xlsx(df: pd.DataFrame, status: str) -> bytes:
    output = io.BytesIO()
    subset = df[df["Overall Status"] == status].copy()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        subset.to_excel(writer, index=False, sheet_name="Results")
        ws = writer.sheets["Results"]
        _style_header(ws, subset)
        _apply_column_widths(ws, subset)
        _color_rows(ws, subset, "Overall Status")
    return output.getvalue()


def dataframe_to_pdf_bytes(df: pd.DataFrame) -> bytes:
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 16)
    pdf.cell(0, 10, "Email Verification Report", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)
    pdf.set_font("Helvetica", "", 10)
    pdf.cell(0, 8, f"Generated: {datetime.now().strftime('%d-%b-%Y %I:%M %p')}", align="C", new_x="LMARGIN", new_y="NEXT")

    total = len(df)
    valid = int((df["Overall Status"] == "VALID").sum())
    risky = int((df["Overall Status"] == "RISKY").sum())
    invalid = int((df["Overall Status"] == "INVALID").sum())

    pdf.ln(5)
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, f"Total: {total}  |  Valid: {valid}  |  Risky: {risky}  |  Invalid: {invalid}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)

    cols = list(df.columns)
    display_cols = [c for c in cols if c not in ("LinkedIn URL", "SMTP Response")]
    if not display_cols:
        display_cols = cols[:8]

    col_width = max(22, int(180 / len(display_cols)))

    pdf.set_font("Helvetica", "B", 6)
    pdf.set_fill_color(68, 114, 196)
    pdf.set_text_color(255, 255, 255)
    for col in display_cols:
        pdf.cell(col_width, 7, col[:18], border=1, align="C", fill=True)
    pdf.ln()
    pdf.set_text_color(0, 0, 0)

    pdf.set_font("Helvetica", "", 5)
    for _, row in df.iterrows():
        overall = str(row.get("Overall Status", ""))
        if overall == "VALID":
            pdf.set_fill_color(212, 237, 222)
        elif overall == "RISKY":
            pdf.set_fill_color(255, 243, 205)
        else:
            pdf.set_fill_color(248, 215, 218)

        if pdf.get_y() > 265:
            pdf.add_page()
            pdf.set_font("Helvetica", "B", 6)
            pdf.set_fill_color(68, 114, 196)
            pdf.set_text_color(255, 255, 255)
            for col in display_cols:
                pdf.cell(col_width, 7, col[:18], border=1, align="C", fill=True)
            pdf.ln()
            pdf.set_text_color(0, 0, 0)
            pdf.set_font("Helvetica", "", 5)

        for col in display_cols:
            val = str(row.get(col, ""))[:22]
            pdf.cell(col_width, 5, val, border=1, align="C", fill=True)
        pdf.ln()

    return bytes(pdf.output(dest="S"))
